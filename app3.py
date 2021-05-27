# módulo para trabalhar com pdf
import pdfplumber as p

# módulo para trabalhar com expressões regulares
import re

# módulo para trabalhar com excel
from openpyxl import *

# abrindo o arquivo pdf existente
nome = input('Insira o nome do pdf do qual deseja extrair os dados(Pressione Ctrl+c para sair)\n>>>>')
with p.open(f'relatorios-exemplos/{nome}') as pdf:
	page = pdf.pages[0] # determinando a página para extrair o texto
	text = page.extract_text() # extraindo o texto



# procurando todas as ocorrências de "Pagador" com 're.findall()
fav = re.findall(r'Favorecido:(.*?) Inscrição', text)

inscricao = re.findall(r'Inscrição: (.*?) Id', text)

# procurando todas as ocorrências de "Valor" com 're.findall()
valor_pag = re.findall(r'Valor Pag.: (.*?) Data Pag', text)

data_pagto = re.findall(r'Data Pag.:(.*?) Nr', text)

banco = re.findall(r'Banco: (.*?) Agência', text)

agencia = re.findall(r'Agência: (.*?) Conta', text)

conta = re.findall(r'Conta:(.*?)\n', text)

mensagem = re.findall(r'Mensagem:(.*?) Nr', text)




# adicionando todos os dados de cada pessoa em sequência para posterior transferência para o excel, pois o openpyxl adiciona ao excel dados em forma de listas para preencher a linha inteira de cada registro.
# neste caso, será gerada uma única lista com todos os dados em sequência.
dados = []
for i in range(len(fav)):
    dados.append(fav[i])
    dados.append(inscricao[i])
    dados.append(valor_pag[i])
    dados.append(data_pagto[i])
    dados.append(banco[i])
    dados.append(agencia[i])
    dados.append(conta[i])
    dados.append(mensagem[i])

# mostrando os dados
# print(dados)


# criação do arquivo xlsx através do openpyxl:
#wb = Workbook()
wb = load_workbook('planilha.xlsx')

# dando o nome plan pra planilha ativa:
plan = wb.active 

# definindo valores de cabeçalho:
plan['A1'] = "Favorecido"
plan['B1'] = "Inscrição"
plan['C1'] = "Valor Pago"
plan['D1'] = 'Data de pagto.'
plan['E1'] = "Banco"
plan['F1'] = "Agência"
plan['G1'] = "Conta"
plan['H1'] = "Mensagem"

################
# função para dividir a lista completa em 5 partes iguais, ou seja, à cada pessoa pertencem 5 itens da lista(pagador, valor, vencimento, beneficiário e data de pagamento). Se dividirmos a lista geral por 5, teremos sublistas com os dados de cada pessoa a serem inseridos à planilha
def dividir_lista(lista, n_elementos):
    # para cada item na lista num alcance de 0 ao número correspondente ao tamanho da lista, num passo determinado(no caso, será 5):
    for i in range(0, len(lista), n_elementos): 
        yield lista[i:i + n_elementos]
  
# quantos elementos cada lista deve ter
n = 8

# x é igual a lista de sublistas gerada pela função acima e receberá como argumentos a lista bruta e o número de sublistas desejado:
x = list(dividir_lista(dados, n))
# print (x)
################

# inserindo na planilha do excel as sublistas dentro da lista 'x' usando um loop.
for i in x:
    plan.append(i)

# salvando planilha
wb.save(filename = 'planilha.xlsx')

print('Dados copiados com sucesso')

#leitura do xlsx
#lwb = load_workbook(filename = 'faturas.xlsx')
