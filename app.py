import pdfplumber as p
import re
from openpyxl import *


with p.open('faturas.pdf') as pdf:
	page = pdf.pages[0]
	text = page.extract_text()

	print(text)

print('===================================================================')
print()
print()
print()
print()
print()
print()


pagador = re.findall(r'(?<=Pagador:\s)[A-Z].*', text)
print(f'Pagador: {pagador}')

valor = re.findall(r'(?<=Valor:\s)\d*\.\d*\,\d{2}', text)
print(f'Valor: {valor}')

venc = re.findall(r'(?<=Vencimento:\s)\d{2}\/\d{2}\/\d{4}', text)
# o excel nao aceitou data com barra nesse caso, entao troquei por hifen
venc_formatado = []
for item in venc:
	troca = item.replace('/', '-')
	venc_formatado.append(troca)
print(f'Vencimento: {li}')

benefic = re.findall(r'(?<=Beneficiário:\s)[A-Z].*', text)
print(f'Beneficiário: {benefic}')

data_hora = re.findall(r'\d{2}/\d{2}/\d{4}\s\d{2}\:\d{2}', text)
print(f'Data: {data_hora}')


# criação do xlsx:
wb = Workbook()

plan = wb.active #dando o nome plan pra planilha ativa

plan['A1'] = "Nome"
plan['B1'] = "Valor"
plan['C1'] = "Vencimento"
plan['D1'] = "Data pagto."


# passando pagador pro excel
j = 0
for p in pagador:
	plan.cell(row=j+2, column=1).value = p
	j = j + 1

# passando valor pro excel
j = 0
for va in valor:
	plan.cell(row=j+2, column=2).value = va
	j = j + 1

# passando vencimento pro excel
j = 0
for ve in venc_formatado:
	plan.cell(row=j+2, column=3).value = ve
	j = j + 1

# passando a data de pagamento pro excel
j = 0
for d in data_hora:
	plan.cell(row=j+2, column=4).value = d
	j = j + 1

# salvando planilha
wb.save(filename = 'faturas.xlsx')

#leitura do xlsx
lwb = load_workbook(filename = 'faturas.xlsx')